from sys import exit
from typing import NewType, Any
import json
import string


from bs4 import BeautifulSoup
from requests import Session, Response
import openpyxl
import requests

# ANNOTATIONS
url_str = NewType('url_str', str)
# CONSTANTS
USER_AGENT = 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:81.0) Gecko/20100101 Firefox/81.0'
HOST: url_str = 'https://www.rusbonds.ru'
COLUMNS_NAMES: list[str] = [
    '№',
    'Дата окончания купона',
    'Период, дней',
    'Ставка, % год.',
    'Сумма на 1 обл, RUB',
    'Примечание'
]
with open('login_data.json') as login_data_json:
    DATA_FOR_LOGGING: dict[str: str] = json.load(open('login_data.json'))


def login_user_on_site() -> Session:
    """Login in site and return session with logging user."""
    path_to_login_page = '/auth.asp'
    registered_user: Session = requests.Session()

    registered_user.headers['User-Agent'] = USER_AGENT
    registered_user.post(HOST + path_to_login_page, data=DATA_FOR_LOGGING)

    return registered_user


def write_coupon_data_in_excel(row_data: list[Any]) -> None:
    """
    Write row data in coupon_data.xlsx file. If this file not exits,
    function to create this file and write column name.

    row_data: list[Any] -> list consist of any data.
    """
    xlsx_file_name: str = 'coupon_data.xlsx'

    try:
        coupon_data_workbook = openpyxl.load_workbook(xlsx_file_name)
    except FileNotFoundError:
        new_coupon_data_workbook = openpyxl.Workbook()

        # Delete default list in workbook.
        for sheet_name in new_coupon_data_workbook.sheetnames:
            sheet = new_coupon_data_workbook[sheet_name]
            new_coupon_data_workbook.remove(sheet)

        coupon_data_list = new_coupon_data_workbook.create_sheet('Coupon data')

        for column in range(1, len(COLUMNS_NAMES) + 1):
            coupon_data_list.cell(row=1, column=column).value = COLUMNS_NAMES[column - 1]

        new_coupon_data_workbook.save(xlsx_file_name)
        new_coupon_data_workbook.close()

        coupon_data_workbook = openpyxl.load_workbook(xlsx_file_name)

    coupon_data_list = coupon_data_workbook['Coupon data']

    row: int = 2
    column_letters = string.ascii_uppercase[:len(row_data)]
    row_is_free: bool = False

    while True:

        for col_letter in column_letters:
            if coupon_data_list[col_letter + str(row)].value is None:
                row_is_free = True
            else:
                row_is_free = False
                break

        if row_is_free:
            for column in range(1, len(row_data) + 1):
                coupon_data_list.cell(row=row, column=column).value = row_data[column - 1]
            row += 1
            break
        else:
            row += 1
    try:
        coupon_data_workbook.save(xlsx_file_name)
        coupon_data_workbook.close()
    except PermissionError:
        print('Пожалуйста, закройте файл coupon_data.xlsx, так как парсер не может работать с открытым файлом для вывода.')
        exit()


def parse_coupon_data():
    """
    Parsed table from https://www.rusbonds.ru/emit_coup.asp?tool=168072 site by name 'Список купонных выплат'.
    If user not logging in this site, print in console that user is not registered.
    """
    logging_user: Session = login_user_on_site()

    path_to_page_with_data_for_parser = '/emit_coup.asp?tool=168072'
    data_in_row: list[Any] = []

    page_with_data_for_parser: Response = logging_user.get(HOST + path_to_page_with_data_for_parser)
    bs_page_with_data_for_parser: BeautifulSoup = BeautifulSoup(page_with_data_for_parser.content, 'html.parser')
    if ('Информация доступна только зарегистрированным пользователям.' not in
            bs_page_with_data_for_parser.get_text()):
        table_with_coupon_data = bs_page_with_data_for_parser.find('table', class_='tbl_data')

        for row_data in table_with_coupon_data.find('tbody').find_all('tr'):
            for data in row_data.find_all('td'):
                data_in_row.append(data.get_text())
            else:
                write_coupon_data_in_excel(data_in_row)
                data_in_row = []
    else:
        print('Ваш пользователь не зарегистрирован на сайте, либо его данные неверны.')
        exit()


def run():
    parse_coupon_data()
    print('Парсер завершил свою работу.')

if __name__ == '__main__':
    run()